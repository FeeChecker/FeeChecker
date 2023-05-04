from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl

from models import Chain


def get_if_exist(tag: str, di: dict):
    if di is None:
        return None
    if tag in di:
        return di[tag]
    if tag == 'in_usd':
        return 0
    return None


def name_by_chain(chain):
    match chain:
        case 'arb':
            return "Arbitrum"
        case 'eth':
            return "Ethereum"
        case 'opt':
            return "Optimism"
        case 'ftm':
            return "Fantom"
        case 'matic':
            return "Polygon"
        case 'zk':
            return "ZkSync"
        case 'bsc':
            return "BSC"


def create_tables_by_accounts(accounts, out_source, bookName):
    wb = openpyxl.Workbook()

    chains = [
        # Chain.ChainName.ARBITRUM,
        # Chain.ChainName.OPTIMISM,
        Chain.ChainName.ETHEREUM,
        Chain.ChainName.FANTOM,
        Chain.ChainName.POLYGON,
        Chain.ChainName.BSC]

    arbitrum = wb.active
    arbitrum.title = 'ZkSync'
    arbitrum.append(['Wallet', 'Transaction count', 'First trans UTC',
                     'Last trans UTC', 'AVG trans time', 'Fee(token)', 'Fee(USD)'])
    for cell in arbitrum['1']:
        cell.font = Font(color='ffffff', b=True, size=14)
        cell.fill = PatternFill('solid', fgColor='000000')
    chain_name = 'zk'
    line = 2
    side = Side(border_style="medium", color="000000")
    for ind, acc in enumerate(accounts):
        for fee_token in get_if_exist(chain_name, acc.fee_pay):
            first_token = list(get_if_exist(chain_name, acc.fee_pay))[0]
            fee_tk = acc.fee_pay[chain_name][fee_token]
            if fee_token == first_token:
                arbitrum.append([acc.wallet,
                                 get_if_exist(chain_name, acc.trans_count),
                                 get_if_exist(chain_name, acc.fist_trans),
                                 get_if_exist(chain_name, acc.last_trans),
                                 get_if_exist(chain_name, acc.avg_time_trans),
                                 f"{fee_tk['count']} ( {fee_tk['symbol']} )",
                                 fee_tk['in_usd']])
            else:
                arbitrum.append([None,
                                 None,
                                 None,
                                 None,
                                 None,
                                 f"{fee_tk['count']} ( {fee_tk['symbol']} )",
                                 fee_tk['in_usd']])

            for cell in arbitrum[str(line)]:
                if ind % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='d8d8d8')
                else:
                    cell.fill = PatternFill('solid', fgColor='bfbfbf')
                cell.border = Border(right=side)
            print(ind, line)
            line += 1

    arbitrum.merge_cells("K1:O1")
    arbitrum['K1'].font = Font(b=True, size=14)
    arbitrum['K1'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K1'] = 'ETH SPENT ON GAS'
    arbitrum.merge_cells('K2:O7')
    arbitrum['K2'].font = Font(b=True, size=20)
    arbitrum['K2'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K2'].fill = PatternFill('solid', fgColor='f2f2f2')
    arbitrum['K2'].border = Border(
        right=side, top=side, left=side, bottom=side)
    arbitrum['K2'] = '=SUM(F:F)'

    arbitrum.merge_cells("K9:O9")
    arbitrum['K9'].font = Font(b=True, size=14)
    arbitrum['K9'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K9'] = '$USD SPENT ON GAS (at current price)'
    arbitrum.merge_cells('K10:O15')
    arbitrum['K10'].font = Font(b=True, size=20)
    arbitrum['K10'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K10'].fill = PatternFill('solid', fgColor='f2f2f2')
    arbitrum['K10'].border = Border(
        right=side, top=side, left=side, bottom=side)
    arbitrum['K10'] = '=SUM(G:G)'

    arbitrum.column_dimensions['A'].width = 45
    arbitrum.column_dimensions['B'].width = 20
    arbitrum.column_dimensions['C'].width = 20
    arbitrum.column_dimensions['D'].width = 20
    arbitrum.column_dimensions['E'].width = 17
    arbitrum.column_dimensions['F'].width = 25
    arbitrum.column_dimensions['G'].width = 15


    # arbitrumTokens = wb.create_sheet(name_by_chain(chain_name) + ' tokens')
    # for acc in accounts:
    #   arbitrumTokens.append([acc.wallet, 'Tokens', 'USD', '===NEW WALLET==='])
    #   tokens = get_if_exist(chain_name, acc.token_balance)
    #   if tokens is None:
    #       continue
    #   for token_name in tokens:
    #       tk = tokens[token_name]
    #       arbitrumTokens.append([f"{token_name}", tk['count'], get_if_exist('in_usd', tk)])

    # arbitrumTokens.column_dimensions['A'].width = 45
    # arbitrumTokens.column_dimensions['B'].width = 20
    # arbitrumTokens.column_dimensions['C'].width = 20

    # arbitrumNFTs = wb.create_sheet('Arbitrum NFTs')
    # for acc in accounts:
    #   arbitrumNFTs.append([acc.wallet, 'Name', 'Count', 'USD', '===NEW WALLET==='])
    #   chain_nfts = get_if_exist(chain_name, acc.nfts)
    #   if chain_nfts is None:
    #       continue
    #   for nft_name in chain_nfts:
    #       arbitrumNFTs.append([None, nft_name, chain_nfts[nft_name]['count'], get_if_exist('in_usd', chain_nfts[nft_name])])

    # SHITLY OUT_SOURCE
    for chain_name in out_source:
        arbitrum = wb.create_sheet(name_by_chain(chain_name))
        arbitrum.title = chain_name.capitalize()
        arbitrum.append(['Wallet', 'Transaction count', 'First trans UTC',
                         'Last trans UTC', 'AVG trans time', 'Fee(token)', 'Fee(USD)'])
        for cell in arbitrum['1']:
            cell.font = openpyxl.styles.fonts.Font(
                color='ffffff', b=True, size=14)
            cell.fill = PatternFill('solid', fgColor='000000')
        line = 2

        for acc in out_source[chain_name]:
            if acc[0]:
                wal = acc[-1]
                acc = acc[:-1]
                acc.insert(0, wal)
                arbitrum.append(acc)
            else:
                arbitrum.append([acc[1]])

            for cell in arbitrum[str(line)]:
                if ind % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='d8d8d8')
                else:
                    cell.fill = PatternFill('solid', fgColor='bfbfbf')
                cell.border = Border(right=side)
            line += 1

        arbitrum.merge_cells("K1:O1")
        arbitrum['K1'].font = Font(b=True, size=14)
        arbitrum['K1'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K1'] = 'ETH SPENT ON GAS'
        arbitrum.merge_cells('K2:O7')
        arbitrum['K2'].font = Font(b=True, size=20)
        arbitrum['K2'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K2'].fill = PatternFill('solid', fgColor='f2f2f2')
        arbitrum['K2'].border = Border(
            right=side, top=side, left=side, bottom=side)
        arbitrum['K2'] = '=SUM(F:F)'

        arbitrum.merge_cells("K9:O9")
        arbitrum['K9'].font = Font(b=True, size=14)
        arbitrum['K9'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K9'] = '$USD SPENT ON GAS (at current price)'
        arbitrum.merge_cells('K10:O15')
        arbitrum['K10'].font = Font(b=True, size=20)
        arbitrum['K10'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K10'].fill = PatternFill('solid', fgColor='f2f2f2')
        arbitrum['K10'].border = Border(
            right=side, top=side, left=side, bottom=side)
        arbitrum['K10'] = '=SUM(G:G)'
        
        arbitrum.column_dimensions['A'].width = 45
        arbitrum.column_dimensions['B'].width = 20
        arbitrum.column_dimensions['C'].width = 20
        arbitrum.column_dimensions['D'].width = 20
        arbitrum.column_dimensions['E'].width = 17
        arbitrum.column_dimensions['F'].width = 25
        arbitrum.column_dimensions['G'].width = 15

    for chain_name in chains:
        arbitrum = wb.create_sheet(name_by_chain(chain_name))
        arbitrum.append(['Wallet', 'Transaction count', 'First trans UTC',
                         'Last trans UTC', 'AVG trans time', 'Fee(token)', 'Fee(USD)'])
        for cell in arbitrum['1']:
            cell.font = openpyxl.styles.fonts.Font(
                color='ffffff', b=True, size=14)
            cell.fill = PatternFill('solid', fgColor='000000')
        line = 2
        for ind, acc in enumerate(accounts):
            arbitrum.append([acc.wallet,
                             get_if_exist(chain_name, acc.trans_count),
                             get_if_exist(chain_name, acc.fist_trans),
                             get_if_exist(chain_name, acc.last_trans),
                             get_if_exist(chain_name, acc.avg_time_trans),
                             get_if_exist('count', get_if_exist(
                                 chain_name, acc.fee_pay)),
                             get_if_exist('in_usd', get_if_exist(chain_name, acc.fee_pay))])

            for cell in arbitrum[str(line)]:
                if ind % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='d8d8d8')
                else:
                    cell.fill = PatternFill('solid', fgColor='bfbfbf')
                cell.border = Border(right=side)
            line += 1

        arbitrum.merge_cells("K1:O1")
        arbitrum['K1'].font = Font(b=True, size=14)
        arbitrum['K1'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K1'] = 'ETH SPENT ON GAS'
        arbitrum.merge_cells('K2:O7')
        arbitrum['K2'].font = Font(b=True, size=20)
        arbitrum['K2'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K2'].fill = PatternFill('solid', fgColor='f2f2f2')
        arbitrum['K2'].border = Border(
            right=side, top=side, left=side, bottom=side)
        arbitrum['K2'] = '=SUM(F:F)'

        arbitrum.merge_cells("K9:O9")
        arbitrum['K9'].font = Font(b=True, size=14)
        arbitrum['K9'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K9'] = '$USD SPENT ON GAS (at current price)'
        arbitrum.merge_cells('K10:O15')
        arbitrum['K10'].font = Font(b=True, size=20)
        arbitrum['K10'].alignment = Alignment(
            horizontal='center', vertical='center')
        arbitrum['K10'].fill = PatternFill('solid', fgColor='f2f2f2')
        arbitrum['K10'].border = Border(
            right=side, top=side, left=side, bottom=side)
        arbitrum['K10'] = '=SUM(G:G)'
        
        arbitrum.column_dimensions['A'].width = 45
        arbitrum.column_dimensions['B'].width = 20
        arbitrum.column_dimensions['C'].width = 20
        arbitrum.column_dimensions['D'].width = 20
        arbitrum.column_dimensions['E'].width = 17
        arbitrum.column_dimensions['F'].width = 25
        arbitrum.column_dimensions['G'].width = 15

        # ====== TOKENS OUTPUT PART ======

        # arbitrumTokens = wb.create_sheet(name_by_chain(chain_name) + ' tokens')
        # for acc in accounts:
        #   arbitrumTokens.append([acc.wallet, 'Tokens', 'USD', '===NEW WALLET==='])
        #   tokens = get_if_exist(chain_name, acc.token_balance)
        #   if tokens is None:
        #       continue
        #   for token_name in tokens:
        #       tk = tokens[token_name]
        #       arbitrumTokens.append([f"{token_name} ( {get_if_exist('symbol', tk)} )", tk['count'], get_if_exist('in_usd', tk)])

        # arbitrumTokens.column_dimensions['A'].width = 45
        # arbitrumTokens.column_dimensions['B'].width = 20
        # arbitrumTokens.column_dimensions['C'].width = 20

        # arbitrumNFTs = wb.create_sheet(name_by_chain(chain_name) + ' NFTs')
        # for acc in accounts:
        #   arbitrumNFTs.append([acc.wallet, 'Name', 'Count', 'USD', '===NEW WALLET==='])
        #   chain_nfts = get_if_exist(chain_name, acc.nfts)
        #   if chain_nfts is None:
        #       continue
        #   for nft_name in chain_nfts:
        #       arbitrumNFTs.append([None, nft_name, chain_nfts[nft_name]['count'], get_if_exist('in_usd', chain_nfts[nft_name])])

    wb.save(str(bookName) + '.xlsx')
